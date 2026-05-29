#!/usr/bin/env python3
"""Lステップ Excel → data/metadata/lstep_progress.csv（特進フィルタ）。"""
from __future__ import annotations

import argparse
import csv
import re
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Downloads" / "Lステップの顧客データ260525.xlsx"
OUTPUT_PATH = ROOT / "data" / "metadata" / "lstep_progress.csv"

SHEET_PROGRAM = "投稿プログラム（新）"
SHEET_POSTS = "投稿数"
STEP_COLS = [f"STEP{i}完了日" for i in range(1, 20)]
POST_COL_RE = re.compile(r"^(\d+)投稿目完了日$")


def parse_date(val) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    if not s or s.lower() in ("nat", "none"):
        return None
    m = re.match(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s.replace(".", "/"))
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def norm_id(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def is_tokushin(course: str | None) -> bool:
    return "特進" in (course or "")


def row_to_map(headers: list, row: tuple) -> dict:
    out: dict = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        key = str(h).strip()
        if not key:
            continue
        out[key] = row[i] if i < len(row) else None
    return out


def steps_summary(row: dict) -> tuple[int, int | None, str | None]:
    """完了STEP数・最新STEP番号・最新STEP完了日。"""
    latest_n = 0
    latest_date = None
    for i, col in enumerate(STEP_COLS, start=1):
        d = parse_date(row.get(col))
        if d:
            latest_n = i
            latest_date = d
    return latest_n, latest_n or None, latest_date


def count_posts(row: dict) -> tuple[str | None, int]:
    first = None
    count = 0
    for key, val in row.items():
        m = POST_COL_RE.match(key or "")
        if not m:
            continue
        d = parse_date(val)
        if d:
            count += 1
            if m.group(1) == "1":
                first = d
    return first, count


def read_program_rows(path: Path, tokushin_only: bool) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_PROGRAM not in wb.sheetnames:
        raise ValueError(f"シート '{SHEET_PROGRAM}' がありません: {path}")
    ws = wb[SHEET_PROGRAM]
    headers: list | None = None
    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        if not headers:
            break
        data = row_to_map(headers, row)
        name = (data.get("表示名") or "").strip()
        if not name:
            continue
        course = (data.get("コース") or "").strip()
        if tokushin_only and not is_tokushin(course):
            continue
        steps_done, latest_step, latest_step_date = steps_summary(data)
        rows.append(
            {
                "management_id": norm_id(data.get("管理ID")),
                "display_name": name,
                "mg_name_lstep": (data.get("担当MG名") or "").strip(),
                "class_name_lstep": (data.get("クラス名(講師名)") or "").strip(),
                "course": course,
                "program_start_date": parse_date(data.get("投稿プログラム開始日")),
                "join_form_date": parse_date(data.get("入会フォーム回答日")),
                "steps_completed": steps_done,
                "latest_step": latest_step or "",
                "latest_step_date": latest_step_date or "",
                "step1_date": parse_date(data.get("STEP1完了日")) or "",
            }
        )
    wb.close()
    return rows


def read_posts_index(path: Path) -> dict[str, dict]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_POSTS not in wb.sheetnames:
        raise ValueError(f"シート '{SHEET_POSTS}' がありません: {path}")
    ws = wb[SHEET_POSTS]
    headers: list | None = None
    by_id: dict[str, list[dict]] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        data = row_to_map(headers, row)
        mid = norm_id(data.get("管理ID"))
        if not mid:
            continue
        first_post, post_count = count_posts(data)
        entry = {
            "first_post_date": first_post or "",
            "post_count": post_count,
            "course_posts": (data.get("コース") or "").strip(),
        }
        by_id.setdefault(mid, []).append(entry)

    merged: dict[str, dict] = {}
    for mid, entries in by_id.items():
        tokushin_rows = [e for e in entries if is_tokushin(e.get("course_posts"))]
        pool = tokushin_rows or entries
        best = max(pool, key=lambda e: (e["post_count"], e["first_post_date"] or ""))
        merged[mid] = best
    wb.close()
    return merged


def effective_sp_start(row: dict) -> str:
    return (
        row.get("program_start_date")
        or row.get("step1_date")
        or row.get("join_form_date")
        or ""
    )


def import_lstep(
    xlsx_path: Path,
    output_path: Path = OUTPUT_PATH,
    *,
    tokushin_only: bool = True,
) -> Path:
    program = read_program_rows(xlsx_path, tokushin_only=tokushin_only)
    posts = read_posts_index(xlsx_path)
    imported_at = datetime.now().isoformat(timespec="seconds")

    fieldnames = [
        "management_id",
        "display_name",
        "mg_name_lstep",
        "class_name_lstep",
        "course",
        "program_start_date",
        "join_form_date",
        "step1_date",
        "sp_start_lstep",
        "steps_completed",
        "latest_step",
        "latest_step_date",
        "first_post_date",
        "post_count",
        "imported_at",
        "source_file",
    ]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in program:
            mid = row["management_id"]
            post = posts.get(mid, {})
            first_post = post.get("first_post_date", "")
            post_count = post.get("post_count", 0)
            sp = effective_sp_start(
                {
                    "program_start_date": row.get("program_start_date") or "",
                    "step1_date": row.get("step1_date") or "",
                    "join_form_date": row.get("join_form_date") or "",
                }
            )
            w.writerow(
                {
                    **row,
                    "sp_start_lstep": sp,
                    "first_post_date": first_post,
                    "post_count": post_count,
                    "imported_at": imported_at,
                    "source_file": xlsx_path.name,
                }
            )

    return output_path


def main() -> None:
    p = argparse.ArgumentParser(description="Lステップ Excel を CSV に取込")
    p.add_argument(
        "xlsx",
        nargs="?",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"入力 xlsx（既定: {DEFAULT_XLSX}）",
    )
    p.add_argument("-o", "--output", type=Path, default=OUTPUT_PATH)
    p.add_argument(
        "--all-courses",
        action="store_true",
        help="特進以外も含める（デフォルトは特進のみ）",
    )
    args = p.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"ファイルがありません: {args.xlsx}")

    out = import_lstep(args.xlsx, args.output, tokushin_only=not args.all_courses)
    with out.open(encoding="utf-8") as f:
        n = sum(1 for _ in f) - 1
    print(f"取込完了: {out} ({n} 行)")
    print("次: python scripts/build_dashboard.py")


if __name__ == "__main__":
    main()
