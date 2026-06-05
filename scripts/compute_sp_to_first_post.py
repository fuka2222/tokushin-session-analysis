#!/usr/bin/env python3
"""SP開始日マスタ × Lステップ → SP開始から初投稿完了までの日数。"""

from __future__ import annotations

import argparse
import csv
import html as html_mod
import re
import sys
from datetime import date
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from import_lstep import count_posts, is_tokushin, parse_date as parse_date_str  # noqa: E402
from sp_start_lookup import format_roster, load_sp_start_index, lookup_sp_start, norm_name  # noqa: E402

DEFAULT_LSTEP = Path.home() / "Downloads" / "Lステップの顧客データ 2026年6月3日集計.xlsx"
DEFAULT_SP = ROOT / "data/metadata/sp_start_dates.tsv"
DEFAULT_ROSTER = ROOT / "data/metadata/roster_paste.tsv"
DEFAULT_OUT = ROOT / "data/reports/monthly_20260604/first_post_days_from_sp.csv"


def parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, pd.Timestamp):
        return val if val.year >= 2020 else None
    if isinstance(val, pd.Timestamp):
        if pd.isna(val):
            return None
        d = val.date()
        return d if d.year >= 2020 else None
    s = parse_date_str(val)
    if not s or s.lower() in ("nat", "none"):
        return None
    y, m, d = map(int, s.split("-"))
    return date(y, m, d)


def load_lstep_step18(path: Path) -> dict[str, dict]:
    df = pd.read_excel(path, sheet_name="投稿プログラム（新）")
    df = df[df["表示名"].notna() & (df["表示名"].astype(str).str.strip() != "")]
    index: dict[str, dict] = {}
    for _, r in df.iterrows():
        if not is_tokushin(str(r.get("コース") or "")):
            continue
        name = str(r["表示名"]).strip()
        step18 = parse_date(r.get("STEP18完了日"))
        step1 = parse_date(r.get("STEP1完了日"))
        key = norm_name(name)
        row = {
            "display_name": name,
            "course": str(r.get("コース") or ""),
            "mg_lstep": str(r.get("担当MG名") or ""),
            "step1_date": step1,
            "step18_date": step18,
            "latest_step": 0,
        }
        for i in range(19, 0, -1):
            d = parse_date(r.get(f"STEP{i}完了日"))
            if d:
                row["latest_step"] = i
                break
        existing = index.get(key)
        if not existing or (row["latest_step"], step18 or date.min) > (
            existing["latest_step"],
            existing["step18_date"] or date.min,
        ):
            index[key] = row
    return index


def norm_id(val) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, float):
        if pd.isna(val):
            return ""
        if val == int(val):
            return str(int(val))
    s = str(val).strip()
    if s.lower() in ("nan", "none"):
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    return s


def load_first_post_dates(path: Path) -> dict[str, date]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "投稿数" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["投稿数"]
    headers: list | None = None
    by_id: dict[str, date] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h else "" for h in row]
            continue
        data = {headers[j]: row[j] if j < len(row) else None for j in range(len(headers))}
        mid = norm_id(data.get("管理ID"))
        if not mid:
            continue
        first, _count = count_posts(data)
        if first:
            d = parse_date(first)
            if d:
                by_id[mid] = d
    wb.close()

    df = pd.read_excel(path, sheet_name="投稿プログラム（新）")
    id_to_name: dict[str, str] = {}
    for _, r in df.iterrows():
        mid = norm_id(r.get("管理ID"))
        name = str(r.get("表示名") or "").strip()
        if mid and name and is_tokushin(str(r.get("コース") or "")):
            id_to_name[mid] = name

    out: dict[str, date] = {}
    for mid, d in by_id.items():
        name = id_to_name.get(mid)
        if name:
            out[norm_name(name)] = d
    return out


def find_lstep(name: str, index: dict[str, dict]) -> dict | None:
    key = norm_name(name)
    if key in index:
        return index[key]
    for k, v in index.items():
        if len(key) >= 3 and len(k) >= 3 and (key in k or k in key):
            return v
    return None


def load_sp_sources(sp_path: Path, roster_path: Path) -> dict[str, date]:
    """SP開始日マスタ優先、ロスターで補完。"""
    index = load_sp_start_index(sp_path)
    if not roster_path.exists():
        return index
    with roster_path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f, delimiter="\t"):
            name = (row.get("生徒名") or "").strip()
            if not name or name.startswith("【") or "テスト" in name:
                continue
            key = norm_name(name)
            if key in index:
                continue
            d = parse_date(row.get("SP開始日", ""))
            if d:
                index[key] = d
    return index


def compute_rows(
    sp_index: dict[str, date],
    lstep: dict[str, dict],
    post1: dict[str, date],
    *,
    sp_month: int | None = None,
    sp_year: int | None = None,
) -> list[dict]:
    rows: list[dict] = []
    for sp_name, sp_start in sorted(sp_index.items(), key=lambda x: x[1]):
        if sp_year and sp_start.year != sp_year:
            continue
        if sp_month and sp_start.month != sp_month:
            continue
        ls = find_lstep(sp_name, lstep)
        display = ls["display_name"] if ls else sp_name
        step18 = ls["step18_date"] if ls else None
        step1 = ls["step1_date"] if ls else None
        pub1_raw = post1.get(norm_name(display)) or post1.get(sp_name)
        pub1 = pub1_raw if pub1_raw and pub1_raw >= sp_start else None
        pub1_note = ""
        if pub1_raw and pub1_raw < sp_start:
            pub1_note = f"不整合(SP前:{pub1_raw.isoformat()})"

        days_step18 = (step18 - sp_start).days if step18 else None
        days_pub1 = (pub1 - sp_start).days if pub1 else None
        days_step1_step18 = (step18 - step1).days if step18 and step1 else None

        within30_step18 = days_step18 is not None and days_step18 <= 30
        within30_pub1 = days_pub1 is not None and days_pub1 <= 30

        rows.append(
            {
                "生徒名": display,
                "SP開始日": sp_start.isoformat(),
                "STEP1完了日": step1.isoformat() if step1 else "",
                "STEP18完了日": step18.isoformat() if step18 else "",
                "1投稿目完了日": pub1.isoformat() if pub1 else "",
                "SP→STEP18日数": days_step18 if days_step18 is not None else "",
                "SP→1投稿目日数": days_pub1 if days_pub1 is not None else "",
                "STEP1→STEP18日数": days_step1_step18 if days_step1_step18 is not None else "",
                "STEP18_30日以内": "Yes" if within30_step18 else ("No" if step18 else ""),
                "1投稿目_30日以内": "Yes" if within30_pub1 else ("No" if pub1 else ""),
                "最新STEP": ls["latest_step"] if ls else "",
                "LステップMG": ls["mg_lstep"] if ls else "",
                "初投稿完了": "STEP18" if step18 else ("未完了" if ls else "Lステップ未照合"),
                "備考": pub1_note,
            }
        )
    return rows


def summary_stats(rows: list[dict], col: str) -> dict:
    vals = [int(r[col]) for r in rows if r.get(col) != ""]
    completed = len(vals)
    total = len(rows)
    within30 = sum(1 for v in vals if v <= 30)
    return {
        "total": total,
        "completed": completed,
        "within30": within30,
        "within30_pct": round(within30 / total * 100, 1) if total else 0,
        "completion_pct": round(completed / total * 100, 1) if total else 0,
        "median": sorted(vals)[len(vals) // 2] if vals else None,
        "mean": round(sum(vals) / len(vals), 1) if vals else None,
        "min": min(vals) if vals else None,
        "max": max(vals) if vals else None,
    }


def write_html(rows: list[dict], stats: dict, out_path: Path, meta: str) -> None:
    def esc(s):
        return html_mod.escape(str(s))

    trs = []
    for r in rows:
        d18 = r.get("SP→STEP18日数", "")
        cls = ""
        if d18 != "":
            cls = ' class="ok"' if int(d18) <= 30 else ' class="ng"'
        trs.append(
            f"<tr{cls}><td>{esc(r['生徒名'])}</td>"
            f"<td>{esc(r['SP開始日'])}</td>"
            f"<td>{esc(r['STEP18完了日'])}</td>"
            f"<td>{esc(d18)}</td>"
            f"<td>{esc(r['STEP18_30日以内'])}</td>"
            f"<td>{esc(r['1投稿目完了日'])}</td>"
            f"<td>{esc(r['SP→1投稿目日数'])}</td>"
            f"<td>{esc(r['初投稿完了'])}</td></tr>"
        )

    s = stats["step18"]
    body = f"""<!DOCTYPE html>
<html lang="ja"><head><meta charset="utf-8">
<title>SP開始→初投稿完了日数</title>
<style>
body{{font-family:system-ui,sans-serif;margin:24px;max-width:1100px}}
table{{border-collapse:collapse;width:100%;font-size:13px}}
th,td{{border:1px solid #ddd;padding:6px 8px;text-align:left}}
th{{background:#f5f5f5}}
.ok{{background:#e8f5e9}} .ng{{background:#fff8e1}}
.stats{{display:flex;gap:24px;flex-wrap:wrap;margin:16px 0}}
.stat{{background:#f9f9f9;padding:12px 16px;border-radius:8px}}
</style></head><body>
<h1>SP開始日 → 初投稿完了（STEP18）の日数</h1>
<p>{esc(meta)}</p>
<div class="stats">
<div class="stat"><strong>対象</strong><br>{s['total']}名（SP開始日マスタ）</div>
<div class="stat"><strong>STEP18完了</strong><br>{s['completed']}名（{s['completion_pct']}%）</div>
<div class="stat"><strong>30日以内</strong><br>{s['within30']}名（{s['within30_pct']}%）</div>
<div class="stat"><strong>中央値</strong><br>{s['median'] if s['median'] is not None else '—'}日</div>
<div class="stat"><strong>平均</strong><br>{s['mean'] if s['mean'] is not None else '—'}日</div>
</div>
<h2>1投稿目（投稿数シート）— 参考</h2>
<p>完了 {stats['pub1']['completed']}名 / 30日以内 {stats['pub1']['within30']}名（{stats['pub1']['within30_pct']}%）</p>
<table>
<thead><tr>
<th>生徒名</th><th>SP開始日</th><th>STEP18完了</th><th>SP→STEP18</th><th>30日以内</th>
<th>1投稿目</th><th>SP→1投稿目</th><th>状態</th>
</tr></thead>
<tbody>
{"".join(trs)}
</tbody></table>
<p style="color:#666;font-size:12px">緑=30日以内 / 黄=31日以上。SP開始日=sp_start_dates.tsv（セッション実施状況管理）</p>
</body></html>"""
    out_path.write_text(body, encoding="utf-8")


def write_md(rows: list[dict], stats: dict, out_path: Path, meta: str) -> None:
    s = stats["step18"]
    p = stats["pub1"]
    lines = [
        "# SP開始日 → 初投稿完了日数",
        "",
        meta,
        "",
        "## サマリ",
        "",
        "### 初投稿作成完了（STEP18）",
        "",
        f"| 指標 | 値 |",
        f"|------|-----|",
        f"| 対象 | **{s['total']}名** |",
        f"| STEP18完了 | **{s['completed']}名**（{s['completion_pct']}%） |",
        f"| SP開始から30日以内 | **{s['within30']}名**（{s['within30_pct']}%） |",
    ]
    if s["median"] is not None:
        lines.append(f"| 中央値（完了者） | {s['median']}日 |")
        lines.append(f"| 平均（完了者） | {s['mean']}日 |")
    lines.extend(
        [
            "",
            "### 実投稿（1投稿目完了日・参考）",
            "",
            f"| 指標 | 値 |",
            f"|------|-----|",
            f"| 1投稿目記録あり | **{p['completed']}名**（{p['completion_pct']}%） |",
            f"| SP開始から30日以内 | **{p['within30']}名**（{p['within30_pct']}%） |",
        ]
    )
    if p["median"] is not None:
        lines.append(f"| 中央値（投稿者） | {p['median']}日 |")
        lines.append(f"| 平均（投稿者） | {p['mean']}日 |")
    lines.extend(
        [
            "",
            "## 定義",
            "",
            "- **SP開始日**: `sp_start_dates.tsv` 優先 → `roster_paste.tsv` で補完",
            "- **初投稿作成完了**: Lステップ新PG `STEP18完了日`",
            "- **実投稿**: 投稿数シート `1投稿目完了日`（SP開始日以降のみ採用）",
            "",
            "## STEP18完了者",
            "",
            "| 生徒名 | SP開始 | STEP18 | SP→STEP18 | 30日以内 | 1投稿目 | SP→1投稿目 |",
            "|--------|--------|--------|-----------|----------|---------|------------|",
        ]
    )
    done = [r for r in rows if r.get("SP→STEP18日数") != ""]
    done.sort(key=lambda r: int(r["SP→STEP18日数"]))
    for r in done:
        lines.append(
            f"| {r['生徒名']} | {r['SP開始日']} | {r['STEP18完了日']} | "
            f"**{r['SP→STEP18日数']}日** | {r['STEP18_30日以内']} | "
            f"{r['1投稿目完了日'] or '—'} | {r['SP→1投稿目日数'] or '—'} |"
        )
    pub_rows = [r for r in rows if r.get("SP→1投稿目日数") != ""]
    pub_rows.sort(key=lambda r: int(r["SP→1投稿目日数"]))
    if pub_rows:
        lines.extend(
            [
                "",
                "## 実投稿あり（1投稿目・日数順）",
                "",
                "| 生徒名 | SP開始 | 1投稿目 | SP→1投稿目 | 30日以内 | 最新STEP |",
                "|--------|--------|---------|------------|----------|----------|",
            ]
        )
        for r in pub_rows:
            lines.append(
                f"| {r['生徒名']} | {r['SP開始日']} | {r['1投稿目完了日']} | "
                f"**{r['SP→1投稿目日数']}日** | {r['1投稿目_30日以内']} | {r['最新STEP']} |"
            )
    pending = [r for r in rows if r.get("SP→STEP18日数") == ""]
    if pending:
        lines.extend(["", "## 未完了 / 未照合", ""])
        for r in pending:
            lines.append(f"- {r['生徒名']}（SP {r['SP開始日']}）— {r['初投稿完了']}")
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    p = argparse.ArgumentParser(description="SP開始→初投稿完了日数")
    p.add_argument("--lstep", type=Path, default=DEFAULT_LSTEP)
    p.add_argument("--sp", type=Path, default=DEFAULT_SP)
    p.add_argument("--roster", type=Path, default=DEFAULT_ROSTER)
    p.add_argument("--out-dir", type=Path, default=ROOT / "data/reports/monthly_20260604")
    p.add_argument("--year", type=int, default=None, help="SP開始年でフィルタ（例: 2026）")
    p.add_argument("--month", type=int, default=None, help="SP開始月でフィルタ（例: 4）")
    args = p.parse_args()

    if not args.lstep.exists():
        print(f"Lステップ not found: {args.lstep}", file=sys.stderr)
        return 1
    if not args.sp.exists():
        print(f"SP master not found: {args.sp}", file=sys.stderr)
        return 1

    sp_index = load_sp_sources(args.sp, args.roster)
    lstep = load_lstep_step18(args.lstep)
    post1 = load_first_post_dates(args.lstep)

    rows = compute_rows(sp_index, lstep, post1, sp_month=args.month, sp_year=args.year)
    stats = {
        "step18": summary_stats(rows, "SP→STEP18日数"),
        "pub1": summary_stats(rows, "SP→1投稿目日数"),
    }

    args.out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = args.out_dir / "first_post_days_from_sp.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()) if rows else [])
        w.writeheader()
        w.writerows(rows)

    filt = ""
    if args.year or args.month:
        filt = f"（SP開始 {args.year or '—'}年{args.month or '—'}月）"
    meta = (
        f"集計: {date.today().isoformat()} / Lステップ: {args.lstep.name} / "
        f"対象 {len(rows)}名 {filt} / SPマスタ+ロスター {len(sp_index)}名"
    )

    write_html(rows, stats, args.out_dir / "first_post_days_from_sp.html", meta)
    write_md(rows, stats, args.out_dir / "first_post_days_from_sp.md", meta)

    s = stats["step18"]
    print(meta)
    print(f"STEP18完了: {s['completed']}/{s['total']} ({s['completion_pct']}%)")
    print(f"30日以内(SP起点): {s['within30']}/{s['total']} ({s['within30_pct']}%)")
    if s["median"] is not None:
        print(f"中央値: {s['median']}日 / 平均: {s['mean']}日")
    print(f"Wrote {csv_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
