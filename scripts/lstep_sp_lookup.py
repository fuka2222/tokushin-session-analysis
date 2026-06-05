#!/usr/bin/env python3
"""Lステップデータから SP開始日（入会フォーム+3日）と SPプログラム完了日・各STEPを引く。"""
from __future__ import annotations

import csv
import re
from datetime import date, datetime, timedelta
from pathlib import Path

STEP_COLS = [f"STEP{i}完了日" for i in range(1, 20)]
SP_OFFSET_DAYS = 3  # 入会フォーム回答日 + 3日 = SP開始日

NAME_ALIASES: dict[str, str] = {
    "くりはらあきこ（ごうだいあきこ）": "ごうだいあきこ",
    "すぎやまももこ": "すぎやま　ももこ",
    "いいじまいつき": "いいじま いつき",
    "たけだみえこ": "たけだ　みえこ",
    "いわさきじゅんこ": "いわさき　じゅんこ",
}


def parse_date(val) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        d = val.date()
        return d if d.year >= 2020 else None
    if isinstance(val, date):
        return val if val.year >= 2020 else None
    s = str(val).strip()
    if not s or s.lower() in ("nat", "none", "-"):
        return None
    m = re.match(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s.replace(".", "/"))
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    try:
        return datetime.fromisoformat(s[:10]).date()
    except ValueError:
        return None


def norm_name(s: str) -> str:
    s = s.strip().lower()
    s = s.replace("\u3000", "").replace(" ", "")
    s = re.sub(r"[（(].*?[）)]", "", s)
    return s


def _record_from_row(
    name: str,
    join: date | None,
    step_dates: list[tuple[int, date]],
    *,
    program_start: date | None = None,
    sp_start_direct: date | None = None,
    step19_date: date | None = None,
) -> dict:
    sp_start = sp_start_direct or program_start or (
        (join + timedelta(days=SP_OFFSET_DAYS)) if join else None
    )
    if not sp_start:
        return {}
    sp_complete = step19_date or (step_dates[-1][1] if step_dates else None)
    return {
        "display_name": name,
        "norm_name": norm_name(name),
        "join_form_date": join,
        "sp_start": sp_start,
        "sp_complete": sp_complete,
        "latest_step": step_dates[-1][0] if step_dates else 0,
        "step_dates": step_dates,
    }


def load_lstep_tsv(path: Path) -> list[dict]:
    rows: list[dict] = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for raw in reader:
            name = (raw.get("表示名") or "").strip()
            if not name:
                continue
            join = parse_date(raw.get("入会フォーム回答日"))
            if not join:
                continue
            step_dates: list[tuple[int, date]] = []
            for i, col in enumerate(STEP_COLS, start=1):
                d = parse_date(raw.get(col))
                if d:
                    step_dates.append((i, d))
            rows.append(_record_from_row(name, join, step_dates))
    return rows


def load_lstep_xlsx(path: Path, *, tokushin_only: bool = True) -> list[dict]:
    from import_lstep import SHEET_PROGRAM, row_to_map

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_PROGRAM not in wb.sheetnames:
        wb.close()
        raise ValueError(f"シート '{SHEET_PROGRAM}' がありません")
    ws = wb[SHEET_PROGRAM]
    headers: list | None = None
    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        data = row_to_map(headers, row)
        name = (data.get("表示名") or "").strip()
        if not name:
            continue
        course = (data.get("コース") or "").strip()
        if tokushin_only and "特進" not in course:
            continue
        join = parse_date(data.get("入会フォーム回答日"))
        program_start = parse_date(data.get("投稿プログラム開始日"))
        step1 = parse_date(data.get("STEP1完了日"))
        sp_direct = program_start or step1
        step_dates: list[tuple[int, date]] = []
        for n, col in enumerate(STEP_COLS, start=1):
            d = parse_date(data.get(col))
            if d:
                step_dates.append((n, d))
        step19 = parse_date(data.get("STEP19完了日"))
        rec = _record_from_row(
            name,
            join,
            step_dates,
            program_start=program_start,
            sp_start_direct=sp_direct,
            step19_date=step19,
        )
        if rec:
            rows.append(rec)
    wb.close()
    return rows


def load_lstep_xlsx_old_sp(path: Path) -> list[dict]:
    """投稿プログラム（旧）の【SP】受講開始日で補完用。"""
    from import_lstep import row_to_map

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "投稿プログラム（旧）" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["投稿プログラム（旧）"]
    headers: list | None = None
    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        data = row_to_map(headers, row)
        name = (data.get("表示名") or "").strip()
        if not name:
            continue
        sp = parse_date(data.get("【SP】受講開始日"))
        if not sp:
            continue
        pp9 = parse_date(data.get("【PP】Step9完了日"))
        rows.append(
            {
                "display_name": name,
                "norm_name": norm_name(name),
                "sp_start": sp,
                "sp_complete": pp9,
                "latest_step": 9 if pp9 else 0,
                "step_dates": [(9, pp9)] if pp9 else [],
                "join_form_date": parse_date(data.get("入会フォーム回答日")),
            }
        )
    wb.close()
    return rows


def _merge_lstep_record(existing: dict | None, new: dict) -> dict:
    if not existing:
        return new
    if (new.get("latest_step") or 0) > (existing.get("latest_step") or 0):
        return new
    if (new.get("latest_step") or 0) == (existing.get("latest_step") or 0):
        if new.get("step_dates") and len(new["step_dates"]) > len(existing.get("step_dates") or []):
            return new
    return existing


def build_lstep_index(path: Path, *, tokushin_only: bool = False) -> dict[str, dict]:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        records = load_lstep_xlsx(path, tokushin_only=tokushin_only)
        for row in load_lstep_xlsx_old_sp(path):
            key = row["norm_name"]
            records.append(row)
    else:
        records = load_lstep_tsv(path)
    index: dict[str, dict] = {}
    for row in records:
        key = row["norm_name"]
        index[key] = _merge_lstep_record(index.get(key), row)
    return index


def lookup_lstep(index: dict[str, dict], commit_name: str) -> dict | None:
    alias = NAME_ALIASES.get(commit_name, commit_name)
    for k in (norm_name(alias), norm_name(commit_name)):
        if k in index:
            return index[k]
    cn = norm_name(commit_name)
    matches = [v for k, v in index.items() if cn in k or k in cn]
    if len(matches) == 1:
        return matches[0]
    return None


def export_lstep_tsv(xlsx_path: Path, out_path: Path) -> Path:
    """xlsx → 貼付更新用 TSV を書き出す。"""
    rows = load_lstep_xlsx(xlsx_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["表示名", "入会フォーム回答日", *STEP_COLS]
    with out_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, delimiter="\t")
        w.writeheader()
        for r in rows:
            row = {"表示名": r["display_name"], "入会フォーム回答日": r["join_form_date"].isoformat()}
            for n, d in r["step_dates"]:
                row[f"STEP{n}完了日"] = d.isoformat()
            w.writerow(row)
    return out_path
